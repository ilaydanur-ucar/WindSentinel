import logging
from pathlib import Path
from typing import List, Dict, Any

from .base import BaseDataSource

logger = logging.getLogger(__name__)


class ExcelDataSource(BaseDataSource):
    """Load data from Excel file using openpyxl."""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)

    def load(self) -> List[Dict[str, Any]]:
        """Load Excel file and return rows as list of dicts."""
        if not self.file_path.exists():
            logger.error(f"Excel file not found: {self.file_path}")
            return []

        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return []

        rows = []
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active

            # First row is header
            headers = [cell.value for cell in ws[1]]

            # Data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: row[i] for i in range(len(headers))}
                rows.append(row_dict)

            logger.info(f"Loaded {len(rows)} rows from Excel: {self.file_path.name}")
        except Exception as e:
            logger.error(f"Error loading Excel: {e}", exc_info=True)

        return rows
